
from gurobipy import *
from numpy import *
from openpyxl import *
from time import *

class AirportPair:
    def __init__(self,origin, destination, distance, demand):
        self.From       = origin
        self.To         = destination
        self.Distance   = distance
        self.Demand     = demand

class Airport:
    def __init__(self,name,hub):
        self.Name   = name
        self.Hub    = hub

class Aircraft:
    def __init__(self,name, cask, lf, seats, speed, lto, bt, fleet, yields, ranges, cost):
        self.Name   = name
        self.CASK   = cask
        self.LF     = lf
        self.Seats  = seats
        self.Speed  = speed
        self.LTO    = lto/60
        self.BT     = bt*7
        self.Fleet  = fleet
        self.Yield  = yields
        self.Range  = ranges
        self.Cost   = cost



def FN_Problem (AirportPairs, Aircrafts, Airports, Budget):
    
    model = Model("FN")                # LP model (this is an object)
    
    w = {}                              # Decision Variables (DVs
    for m in range(len(AirportPairs)):
        w[AirportPairs[m].From,AirportPairs[m].To] = model.addVar(obj=-AirportPairs[m].Distance*Aircrafts[0].Yield, vtype ="I",
                                                name = "w"+AirportPairs[m].From+'-'+AirportPairs[m].To)

    x = {}                              # Decision Variables (DVs)
    for m in range(len(AirportPairs)):
        x[AirportPairs[m].From,AirportPairs[m].To] = model.addVar(obj=-AirportPairs[m].Distance*Aircrafts[0].Yield, vtype ="I",
                                                name = "x"+AirportPairs[m].From+'-'+AirportPairs[m].To)

    z = {}                              # Decision Variables (DVs)
    for k in range(len(Aircrafts)):
        for m in range(len(AirportPairs)):
            z[AirportPairs[m].From,AirportPairs[m].To,k] = model.addVar(obj=Aircrafts[k].CASK*AirportPairs[m].Distance*Aircrafts[k].Seats, vtype ="I",
                                                    name = "z"+AirportPairs[m].From+'-'+AirportPairs[m].To+'-'+Aircrafts[k].Name)

    ACk = {}                              # Decision Variables (DVs)
    for k in range(len(Aircrafts)):
        ACk[k] = model.addVar(obj=0, vtype ="I", name = "AC"+Aircrafts[k].Name)

    model.update()

    Demand = {}                       # build 'capacity' constraints
    for m in range(len(AirportPairs)):
        Demand[AirportPairs[m].From,AirportPairs[m].To] = model.addConstr(x[AirportPairs[m].From,AirportPairs[m].To]+w[AirportPairs[m].From,AirportPairs[m].To],
                                                             '<=', AirportPairs[m].Demand, name = 'qnor'+AirportPairs[m].From+'-'+AirportPairs[m].To)


    # [item for item in accounts if item.get('id')==2]
    # [airp for airp in Airports if airp.get('Name')==AirportPairs[m].From].Name
    # INDX is the airport for which name == AirportPairs[m].From
    DemandTransfer = {}                       # build 'capacity' constraints
    for m in range(len(AirportPairs)):
        DemandTransfer[AirportPairs[m].From,AirportPairs[m].To] = model.addConstr(w[AirportPairs[m].From,AirportPairs[m].To],
                                                             '<=', AirportPairs[m].Demand*next(airp for airp in Airports if airp.Name == AirportPairs[m].From).Hub*next(airp for airp in Airports if airp.Name == AirportPairs[m].To).Hub, 
                                                             name = 'qtrans'+AirportPairs[m].From+'-'+AirportPairs[m].To)

    Capacity = {}                       # build 'capacity' constraints
    for m in range(len(AirportPairs)):
        Capacity[AirportPairs[m].From,AirportPairs[m].To] = model.addConstr(x[AirportPairs[m].From,AirportPairs[m].To]
                                                            + quicksum(w[AirportPairs[m].From,p.Name]*(1-next(airp for airp in Airports if airp.Name == AirportPairs[m].To).Hub) for p in Airports)
                                                            + quicksum(w[p.Name,AirportPairs[m].To]*(1-next(airp for airp in Airports if airp.Name == AirportPairs[m].From).Hub) for p in Airports)
                                                             ,'<=', quicksum(z[AirportPairs[m].From,AirportPairs[m].To,k]*Aircrafts[k].Seats*Aircrafts[k].LF for k in range(len(Aircrafts))), name = 'Capacity'+AirportPairs[m].From+'-'+AirportPairs[m].To)

    Continuity = {}
    for k in range(len(Aircrafts)):                       # build 'continuity' constraints
        for m in range(len(Airports)):    
            Continuity[Airports[m]] = model.addConstr((quicksum(z[Airports[m].Name,p.Name,k] for p in Airports) - quicksum(z[p.Name,Airports[m].Name,k] for p in Airports)),
                                    '=', 0, name ='Continuity' +Airports[m].Name+Aircrafts[k].Name)

    #AirportPairs[m].Distance/Aircrafts[0].Speed+Aircrafts[0].LTO)*z[AirportPairs[m].From,AirportPairs[m].To]

    ACProductivity = {}                       # build 'capacity' constraints
    for k in range(len(Aircrafts)):
        ACProductivity[k] = model.addConstr(quicksum(((AirportPairs[m].Distance/Aircrafts[k].Speed+Aircrafts[k].LTO)*z[AirportPairs[m].From,AirportPairs[m].To,k]) for m in range(len(AirportPairs))),
                                                                '<=', Aircrafts[k].BT*Aircrafts[k].Fleet, name = 'ACProductivity'+Aircrafts[k].Name)

    Range = {}
    for k in range(len(Aircrafts)):
        for m in range(len(AirportPairs)):
            if AirportPairs[m].Distance <= Aircrafts[k].Range:
                a = 10000
            else:
                a = 0
            Range[AirportPairs[m].From,AirportPairs[m].To,k] = model.addConstr(z[AirportPairs[m].From,AirportPairs[m].To,k],
                                                                '<=', a, 
                                                                name = 'range'+AirportPairs[m].From+'-'+AirportPairs[m].To+Aircrafts[k].Name)

    Invest = {}            
    Invest[0] = model.addConstr(quicksum(Aircrafts[k].Cost*ACk[k] for k in range(len(Aircrafts))),
                                                        '<=', Budget, 
                                                        name = 'budget'+AirportPairs[m].From+'-'+AirportPairs[m].To+Aircrafts[k].Name)

    model.update()
    model.write("FN_Model.lp")
    model.optimize()

    status = model.status
    if status != GRB.Status.OPTIMAL:
        if status == GRB.Status.UNBOUNDED:
            print('The model cannot be solved because it is unbounded')
        elif status == GRB.Status.INFEASIBLE:
            print('The model is infeasible; computing IIS')
            model.computeIIS()
            print('\nThe following constraint(s) cannot be satisfied:')
            for c in model.getConstrs():
                if c.IISConstr:
                    print('%s' % c.constrName)
        elif status != GRB.Status.INF_OR_UNBD:
            print('Optimization was stopped with status %d' % status)
        exit(0)

    print
    
    for m in range(len(AirportPairs)):
        print("")
        if x[AirportPairs[m].From,AirportPairs[m].To].X >0 or w[AirportPairs[m].From,AirportPairs[m].To].X >0:
                print("x"+AirportPairs[m].From+'-'+AirportPairs[m].To,x[AirportPairs[m].From,AirportPairs[m].To].X)
                print("w"+AirportPairs[m].From+'-'+AirportPairs[m].To,w[AirportPairs[m].From,AirportPairs[m].To].X)
                
        for k in range(len(Aircrafts)):
            if z[AirportPairs[m].From,AirportPairs[m].To,k].X >0:
                print("z"+AirportPairs[m].From+'-'+AirportPairs[m].To+'-'+Aircrafts[k].Name,z[AirportPairs[m].From,AirportPairs[m].To,k].X)

            
    print
    print ("Objective Function =", model.ObjVal/1.0)
    print ("------------------------------------------------------------------------")

if __name__ == '__main__':
    #=================================================================================================
    # Input excel file with arcs data (sheet1) and commodities data (sheet2)
    AirportPairs    = []
    Aircrafts       = []
    Airports        = []

    
    wb = load_workbook("FNInput.xlsx", read_only=True)
    List_airport_demands = tuple(wb["Demand"].iter_rows())
    List_acparameters = tuple(wb["Aircraft"].iter_rows())
    List_airport_distances = tuple(wb["Distance"].iter_rows())
    List_airports = tuple(wb["Airports"].iter_rows())

    # cask, lf, seats, speed, lto, bt, fleet, yields

    for (name, cask, lf, seats, speed, lto, bt, fleet, yields, ranges, cost) in List_acparameters[1:]:
        new = Aircraft(name.value, float(cask.value),float(lf.value),int(seats.value),int(speed.value),int(lto.value),int(bt.value),int(fleet.value),float(yields.value), int(ranges.value), int(cost.value))
        Aircrafts.append(new)

    for i, origin in enumerate(List_airport_demands[1:]):
        for j, destination in enumerate(origin[1:]):
            new_pair = AirportPair(List_airport_demands[i+1][0].value,List_airport_demands[0][j+1].value,int(List_airport_distances[i+1][j+1].value),int(destination.value))
            AirportPairs.append(new_pair)

    for (name,hub) in List_airports[1:]:
        new = Airport(name.value,int(hub.value))
        Airports.append(new)

    # for pair in Airports:
    #     print(vars(pair))

    # print(next(airp for airp in Airports if airp.Name == AirportPairs[m].From).Hub)

    # print(Airports[[airp for airp in Airports if airp.get('Name')==AirportPairs[0].From].Name].Hub)
  
    del new, new_pair

    

    start_time = time()
    # RUN MCF PROBLEM
    FN_Problem(AirportPairs, Aircrafts, Airports, Budget = 150000000)
    
    elapsed_time = time() - start_time

    print ("Run Time = ", elapsed_time)
    print ("END")
 