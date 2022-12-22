# Group 16
# Mauryze Brug (4700651)
# Alec Pietersz (5020328)
# Emma Zadeits (4671880)

from gurobipy import Model, quicksum, GRB
from numpy import *
from openpyxl import load_workbook
from time import *
import numpy as np
import itertools
import pickle

# file to read input data and generate routes
class AirportPair:
    def __init__(self,origin, destination, distance, demand2030):
        self.From       = origin
        self.To         = destination
        self.Distance   = distance
        self.Demand2030 = demand2030

class DemandAirportPair:
    def __init__(self,origin, destination, demand2020):
        self.From       = origin
        self.To         = destination
        self.Demand2020 = demand2020

class Airport:
    def __init__(self,name, icao, latitude, longitude, runway_length, population2020, population2030, gdp, hub):
        self.Name           = name
        self.ICAO           = icao
        self.Lat            = latitude
        self.Long           = longitude
        self.Runway         = runway_length
        self.Population2020 = population2020
        self.Population2030 = population2030
        self.GDP            = gdp        
        self.Hub            = hub

class Aircraft:
    def __init__(self,name, speed, seats, tat, charging, ac_range, runway, lease_cost, op_cost, time_cost, fuel_cost, energy_cost):
        self.Name           = name
        self.Speed          = speed
        self.Seats          = seats
        self.TAT            = tat/60
        self.Charging       = charging/60
        self.Range          = ac_range
        self.Runway         = runway
        self.Lease          = lease_cost
        self.OperatingCost  = op_cost
        self.TimeCost       = time_cost
        self.FuelCost       = fuel_cost
        self.EnergyCost     = energy_cost
    
class Route:
    def __init__(self, airports, distance, delta, precedent, subsequent, id, legs):
        self.ID         = id
        self.Airports   = airports
        self.Distance   = distance
        self.Delta      = delta
        self.Precedent  = precedent
        self.Subsequent = subsequent        
        self.Legs       = legs

#=================================================================================================
# Input excel file with arcs data (sheet1) and commodities data (sheet2)
fuel_price      = 1.42  # $/gallon
block_time      = 10*7    # hrs
energy_price    = 0.07  # euro/kWh
load_factor     = 0.8
AirportPairs    = []
Aircrafts       = []
Airports        = []
DemandPairs     = []

wbs = load_workbook("Aircraft_info copy.xlsx", read_only=True)
List_aircraft_info = tuple(wbs["Aircraft_info"].iter_rows())

wbs = load_workbook("Group_16_Airport_info copy.xlsx", read_only=True)
List_airport_info = tuple(wbs["Group_16_Airport_info"].iter_rows())

wbs = load_workbook("Group_16_Demand copy.xlsx", read_only=True)
List_demand_forecast_data = tuple(wbs["Group_16_Demand"].iter_rows())

wbs = load_workbook("Group_16_Distances copy.xlsx", read_only=True)
List_airport_distances = tuple(wbs["Group_16_Distances"].iter_rows())

wbs = load_workbook("Group_16_Annual_Growth.xlsx", read_only=True)
List_annual_growth = tuple(wbs["Group_16_Annual_growth"].iter_rows())



annual_growth = float(List_annual_growth[0][0].value)

# origin, destination, distance, demand2020, demand2030
for i, origin in enumerate(List_demand_forecast_data[1:]):
    for j, destination in enumerate(origin[1:]):
        new_pair = DemandAirportPair(List_demand_forecast_data[i+1][0].value,List_demand_forecast_data[0][j+1].value,int(List_demand_forecast_data[i+1][j+1].value))
        DemandPairs.append(new_pair)

# name, icao, latitude, longitude, runway_length, population, gdp, hub
for (name, icao, latitude, longitude, runway_length, population, gdp, hub) in List_airport_info[1:]:
    new = Airport(name.value, icao.value, latitude.value, longitude.value, int(runway_length.value), int(population.value), None, gdp.value, int(hub.value))
    Airports.append(new)

for i, origin in enumerate(List_airport_distances[1:]):
    for j, destination in enumerate(origin[1:]):
        new_pair = AirportPair(List_airport_distances[i+1][0].value,List_airport_distances[0][j+1].value,(List_airport_distances[i+1][j+1].value), None)
        AirportPairs.append(new_pair)


X = np.zeros((len(DemandPairs),4))

for i in range(len(DemandPairs)):
    for j in range(4):
        if j == 0:
            X[i,j] = 1
        elif j == 1:
            # next(airp for airp in Airports if airp.Name == AirportPairs[m].From).Population
            X[i,j] = log(next(airp for airp in Airports if airp.ICAO == DemandPairs[i].From).Population2020*
                        next(airp for airp in Airports if airp.ICAO == DemandPairs[i].To).Population2020)  
        elif j == 2:
            X[i,j] = log(next(airp for airp in Airports if airp.ICAO == DemandPairs[i].From).GDP*
                        next(airp for airp in Airports if airp.ICAO == DemandPairs[i].To).GDP)  
        elif j == 3:
            X[i,j] = log(next(airppair for airppair in AirportPairs if (airppair.From == DemandPairs[i].From and airppair.To == DemandPairs[i].To)).Distance*
                        fuel_price) 

X = np.delete(X, np.where(X < 0)[0], axis=0)

y = np.zeros(len(DemandPairs))
for i in range(len(y)):
    y[i] = log(DemandPairs[i].Demand2020)

y = np.delete(y, np.where(y < 0)[0], axis=0)

solution = np.linalg.lstsq(X,y)[0]

k = solution[0]
b1 = solution[1]
b2 = solution[2]
b3 = solution[3]

# k = -5.908334491674755
# b1 = 0.3507381298874774
# b2 = 0.14095537035529815
# b3 = -0.25329734075161936


for airport in Airports:
    airport.Population2030 = airport.Population2020*(annual_growth**10)

for airport_pair in AirportPairs:
    if airport_pair.From != airport_pair.To:
        population_term = b1*log(next(airp for airp in Airports if (airp.ICAO == airport_pair.From)).Population2030
                                            *next(airp for airp in Airports if (airp.ICAO == airport_pair.To)).Population2030)
        gdp_term = b2*log(next(airp for airp in Airports if (airp.ICAO == airport_pair.From)).GDP
                                            *next(airp for airp in Airports if (airp.ICAO == airport_pair.To)).GDP)
        fuel_term = b3*log(fuel_price*airport_pair.Distance)

        airport_pair.Demand2030 = exp(k + population_term + gdp_term + fuel_term)
    else:
        airport_pair.Demand2030 = 0

for (name, speed, seats, tat, charging, ac_range, runway, lease_cost, op_cost, time_cost, fuel_cost, energy_cost) in List_aircraft_info[1:]:
    new = Aircraft(name.value, speed.value, seats.value, tat.value, charging.value, ac_range.value, runway.value, lease_cost.value, op_cost.value, time_cost.value, fuel_cost.value, energy_cost.value)
    Aircrafts.append(new)

max_range = 0
for acft in Aircrafts:
    if acft.Range > max_range:
        max_range = acft.Range    

hub = next(airp for airp in Airports if (airp.Hub == 0))

temp_list = Airports.copy()
temp_list.remove(hub)

s = []
for apt in temp_list:
    s.append(apt.ICAO)

routes = list(itertools.permutations(s,2))
for apt in temp_list:
    routes.append((apt.ICAO,))

routes2 = []
for route in routes:
    routes2.append((hub.ICAO,) + route + (hub.ICAO,))

valid_routes = []

new_id = 0
for route in routes2:
    range_exceeded = False
    route_length = 0
    for i in range(len(route)-1):
        apt1 = route[i]
        apt2 = route[i+1]
        leg_distance = next(airppair for airppair in AirportPairs if (airppair.From == apt1 and airppair.To == apt2)).Distance
        route_length += leg_distance
        if route_length> max_range:
            range_exceeded = True
    if not range_exceeded and len(route)<=4:
        new = Route(route, route_length, {}, {}, {}, new_id, [])
        valid_routes.append(new)
        new_id+=1

for r, route in enumerate(valid_routes):
    for m in range(len(AirportPairs)):
        if AirportPairs[m].From in route.Airports and AirportPairs[m].To in route.Airports: 
            if route.Airports.index(AirportPairs[m].From) < route.Airports.index(AirportPairs[m].To) and AirportPairs[m].To != hub.ICAO:
                route.Delta[AirportPairs[m].From,AirportPairs[m].To] = 1
            elif AirportPairs[m].To == hub.ICAO and AirportPairs[m].From != hub.ICAO:
                route.Delta[AirportPairs[m].From,AirportPairs[m].To] = 1
            else:
                route.Delta[AirportPairs[m].From,AirportPairs[m].To] = 0
        else:
            route.Delta[AirportPairs[m].From,AirportPairs[m].To] = 0


for r, route in enumerate(valid_routes):
    nodes = {}
    for a in range(len(route.Airports)-1):
        nodes[route.Airports[a]] = route.Airports[a+1:]
    route.Subsequent = nodes  

for r, route in enumerate(valid_routes):
    for a in range(len(route.Airports)-1):
        route.Legs.append((route.Airports[a],route.Airports[a+1]))
        


for r, route in enumerate(valid_routes):
    nodes = {}
    for a in range(len(route.Airports)-1):
        nodes[route.Airports[a]] = route.Airports[a::-1]
    route.Precedent = nodes

routes2 = []
for route in valid_routes:
    if len(route.Airports) > 3:
        routes2.append(route)

routes1 = []
for route in valid_routes:
    if len(route.Airports) <= 3:
        routes1.append(route)

start_time = time()
data = (AirportPairs, Aircrafts, Airports, fuel_price, block_time, energy_price, load_factor, valid_routes, routes2, routes1)
# RUN MCF PROBLEM
with open('routebased_copy.pickle', 'wb') as file:
    pickle.dump(data, file)