# Group 16
# Mauryze Brug (4700651)
# Alec Pietersz (5020328)
# Emma Zadeits (4671880)

from gurobipy import Model, quicksum, GRB
from numpy import *
from openpyxl import load_workbook
from time import *
import numpy as np
import itertools
import pickle
import xlsxwriter
# import datetime
from datetime import datetime, timedelta

class Flight:
    def __init__(self, fn, origin, destination, dep_time, arr_time, ac_cost):
        self.FN             = fn
        self.Origin         = origin
        self.Destination    = destination
        self.DepTime        = dep_time
        self.ArrTime        = arr_time
        self.ACCost         = ac_cost

class Itinerary:
    def __init__(self, id, origin, destination, demand, fare, stops, legs):
        self.ID             = id
        self.Origin         = origin
        self.Destination    = destination
        self.Demand         = demand
        self.Fare           = fare
        self.Stops          = stops
        self.Legs           = legs

class Aircraft:
    def __init__(self, type, units, seats, tat):
        self.Type  = type
        self.Units = units
        self.Seats = seats
        self.TAT   = tat

L      = []
P      = []
N      = set()
K      = []
RR     = {}

Arcs   = {}
Nodes  = {}
NGk    = {}

Delta  = {}


wb = load_workbook("Group_16.xlsx", read_only=True)
List_flights            = tuple(wb["Flight"].iter_rows())
List_itineraries        = tuple(wb["Itinerary"].iter_rows())
List_recapture_rate     = tuple(wb["Recapture Rate"].iter_rows())
List_aircraft           = tuple(wb["Aircraft"].iter_rows())

for flight in List_flights[1:]:
    if flight[0].value:
        ac_cost = {}
        ac_cost[List_flights[0][5].value] = flight[5].value
        ac_cost[List_flights[0][6].value] = flight[6].value
        ac_cost[List_flights[0][7].value] = flight[7].value
        ac_cost[List_flights[0][8].value] = flight[8].value
        ac_cost[List_flights[0][9].value] = flight[9].value
        deptime = datetime.combine(datetime(1970, 1, 1), flight[3].value)
        arrtime = datetime.combine(datetime(1970, 1, 1), flight[4].value)
        new_flight = Flight(flight[0].value,flight[1].value,flight[2].value,deptime,arrtime,ac_cost)
        L.append(new_flight)

P.append(Itinerary(0,'NONE','NONE',100000,0,0,[0]))
for itinerary in List_itineraries[1:]:
    if itinerary[1].value:
        legs = []
        legs.append(itinerary[6].value)
        if itinerary[7].value != 0:
            legs.append(itinerary[7].value)
       
        new_itinerary = Itinerary(itinerary[0].value+1,itinerary[1].value,itinerary[2].value,itinerary[3].value,itinerary[4].value,itinerary[5].value,legs)
        P.append(new_itinerary)


for rr_pair in List_recapture_rate[1:]:
    RR[rr_pair[0].value+1,rr_pair[1].value+1] = rr_pair[2].value

for itin1 in P:
    for itin2 in P:
        if not (itin1.ID,itin2.ID) in RR:
            RR[itin1.ID,itin2.ID] = 0
        if itin1.ID == itin2.ID:
            RR[itin1.ID,itin2.ID] = 1

for itin1 in P:
    RR[itin1.ID,0] = 1

for aircraft in List_aircraft[1:]:
    if aircraft[0].value:
        new_aircraft = Aircraft(aircraft[0].value,aircraft[1].value,aircraft[2].value,aircraft[3].value)
        K.append(new_aircraft)

for flight in L:
    N.add(flight.Origin)
    N.add(flight.Destination)
N = list(N)

for i in L:
    for p in P:
        if i.FN in p.Legs:# or p.ID == 0:
            Delta[i.FN,p.ID] = 1
        else:
            Delta[i.FN,p.ID] = 0

arc_id = 0
for aircraft in K:
    for flight in L:
        temp_time = flight.ArrTime + timedelta(minutes=aircraft.TAT)
        Arcs[arc_id] = {'arc_type':flight.FN,'ac_type':aircraft.Type,'start_ap':flight.Origin,'end_ap':flight.Destination,'start_time':flight.DepTime,'end_time':temp_time.replace(day = 1)}
        arc_id += 1

# node_id = 0
for id, arc in Arcs.items():
    if (arc['ac_type'],arc['start_ap'],arc['start_time']) in Nodes:
        Nodes[arc['ac_type'],arc['start_ap'],arc['start_time']]['o_arcs'].append(id)
    else:
        Nodes[arc['ac_type'],arc['start_ap'],arc['start_time']] = {'o_arcs':[id],'e_arcs':[],'g_o_arc':None,'g_e_arc':None}
        # node_id+=1

    if (arc['ac_type'],arc['end_ap'],arc['end_time']) in Nodes:
        Nodes[arc['ac_type'],arc['end_ap'],arc['end_time']]['e_arcs'].append(id)
    else:
        Nodes[arc['ac_type'],arc['end_ap'],arc['end_time']] = {'o_arcs':[],'e_arcs':[id],'g_o_arc':None,'g_e_arc':None}
        # node_id+=1
# print(Nodes)
    
NNodes = []
for id, node in Nodes.items():
    NNodes.append({'ac_type':id[0],'ap':id[1],'time':id[2],'o_arcs':node['o_arcs'],'e_arcs':node['e_arcs'], 'g_o_arc':node['g_o_arc'], 'g_e_arc':node['g_e_arc']})


for aircraft in K:
    # print(aircraft.Type)
    for airport in N:
        # print(airport)
        temp_nodes = []
        for node in NNodes:
            if node['ac_type'] == aircraft.Type and node['ap'] == airport:
                temp_nodes.append(node['time'])
        # print(temp_nodes)
        sort_list = sorted(temp_nodes)
        # print(sort_list)
        for i in range(len(sort_list)-1):
            Arcs[arc_id] = {'arc_type':'ground','ac_type':aircraft.Type,'start_ap':airport,'end_ap':airport,'start_time':sort_list[i],'end_time':sort_list[i+1]}
            Nodes[aircraft.Type,airport,sort_list[i]]['g_o_arc'] = arc_id
            Nodes[aircraft.Type,airport,sort_list[i+1]]['g_e_arc'] = arc_id
            arc_id += 1            
        Arcs[arc_id] = {'arc_type':'ground','ac_type':aircraft.Type,'start_ap':airport,'end_ap':airport,'start_time':sort_list[-1],'end_time':sort_list[0]}
        Nodes[aircraft.Type,airport,sort_list[-1]]['g_o_arc'] = arc_id
        Nodes[aircraft.Type,airport,sort_list[0]]['g_e_arc'] = arc_id
        arc_id += 1

# print(Arcs)
# print("")
# print(Nodes)

for ac in K:
    cut_arcs = []
    for id, arc in Arcs.items():
        if arc['ac_type'] == ac.Type and (arc['start_time'] > arc['end_time'] or arc['start_time'] == arc['end_time']):
            cut_arcs.append(id)
    NGk[ac.Type] = cut_arcs

# print("")
# print(NGk)

# print(RR)

data = (L,P,N,K,RR,Arcs,Nodes,NGk, Delta)
with open('input_data.pickle', 'wb') as file:
    pickle.dump(data, file)



