#hello
from gurobipy import Model, quicksum, GRB
from numpy import *
from openpyxl import *
from time import *
import numpy as np
import pandas as pd
import xlsxwriter

class AirportPair:
    def __init__(self,origin, destination, distance, demand2030):
        self.From       = origin
        self.To         = destination
        self.Distance   = distance
        self.Demand2030 = demand2030

class DemandAirportPair:
    def __init__(self,origin, destination, demand2020):
        self.From       = origin
        self.To         = destination
        self.Demand2020 = demand2020

class Airport:
    def __init__(self,name, icao, latitude, longitude, runway_length, population2020, population2030, gdp, hub):
        self.Name           = name
        self.ICAO           = icao
        self.Lat            = latitude
        self.Long           = longitude
        self.Runway         = runway_length
        self.Population2020 = population2020
        self.Population2030 = population2030
        self.GDP            = gdp        
        self.Hub            = hub

class Aircraft:
    def __init__(self,name, speed, seats, tat, charging, ac_range, runway, lease_cost, op_cost, time_cost, fuel_cost, energy_cost):
        self.Name           = name
        self.Speed          = speed
        self.Seats          = seats
        self.TAT            = tat/60
        self.Charging       = charging/60
        self.Range          = ac_range
        self.Runway         = runway
        self.Lease          = lease_cost
        self.OperatingCost  = op_cost
        self.TimeCost       = time_cost
        self.FuelCost       = fuel_cost
        self.EnergyCost     = energy_cost


def FN_Problem (AirportPairs, Aircrafts, Airports, fuel_price, block_time, energy_price, load_factor, solve):
    
    model = Model("FN")                # LP model (this is an object)
    
    w = {}                              # Decision Variables (DVs
    for m in range(len(AirportPairs)):
        AirPair = AirportPairs[m]
        if AirPair.Distance > 0:
            yield_rpk = 5.9*AirPair.Distance**-0.76+0.043
        else:
            yield_rpk = 0     
        w[AirPair.From,AirPair.To] = model.addVar(obj=-yield_rpk*AirPair.Distance*0.9, vtype ="I",
                                                name = ("w"+AirPair.From+'-'+AirPair.To).replace(" ", ""))

    x = {}                              # Decision Variables (DVs)
    for m in range(len(AirportPairs)):
        AirPair = AirportPairs[m]
        if AirPair.Distance > 0:
            yield_rpk = 5.9*AirPair.Distance**-0.76+0.043
        else:
            yield_rpk = 0 
        x[AirPair.From,AirPair.To] = model.addVar(obj=-yield_rpk*AirPair.Distance, vtype ="I",
                                                name = ("x"+AirPair.From+'-'+AirPair.To).replace(" ", ""))

    z = {}                              # Decision Variables (DVs)
    for k in range(len(Aircrafts)):
        for m in range(len(AirportPairs)):
            AirPair = AirportPairs[m]  
            acft = Aircrafts[k]
            gi = next(airp for airp in Airports if airp.ICAO == AirPair.From).Hub
            gj = next(airp for airp in Airports if airp.ICAO == AirPair.To).Hub
            z[AirPair.From,AirPair.To,k] = model.addVar(obj=((acft.OperatingCost + acft.TimeCost*AirPair.Distance/acft.Speed + 
                                                                acft.FuelCost*fuel_price/1.5*AirPair.Distance)*(1-(2- gi - gj)*0.3)
                                                                +energy_price*acft.EnergyCost*AirPair.Distance/acft.Range), vtype ="I",
                                                    name = ("z"+ AirPair.From +'-'+AirPair.To+'-'+acft.Name).replace(" ", ""))

    ACk = {}                              # Decision Variables (DVs)
    for k in range(len(Aircrafts)):
        acft = Aircrafts[k]
        ACk[k] = model.addVar(obj=acft.Lease, vtype ="I", name = ("AC"+acft.Name).replace(" ", ""))

    model.update()
    if solve:

        Demand = {}                       # build 'capacity' constraints
        for m in range(len(AirportPairs)):
            AirPair = AirportPairs[m]
            Demand[AirPair.From,AirPair.To] = model.addConstr(x[AirPair.From,AirPair.To]+w[AirPair.From,AirPair.To],
                                                                '<=', AirPair.Demand2030, name = 'qnor'+AirPair.From+'-'+AirPair.To)


        # [item for item in accounts if item.get('id')==2]
        # [airp for airp in Airports if airp.get('Name')==AirportPairs[m].From].Name
        # INDX is the airport for which name == AirportPairs[m].From
        DemandTransfer = {}                       # build 'capacity' constraints
        for m in range(len(AirportPairs)):
            AirPair = AirportPairs[m]
            DemandTransfer[AirPair.From,AirPair.To] = model.addConstr(w[AirPair.From,AirPair.To],
                                                                '<=', AirPair.Demand2030*next(airp for airp in Airports if airp.ICAO == AirPair.From).Hub*next(airp for airp in Airports if airp.ICAO == AirPair.To).Hub, 
                                                                name = 'qtrans'+AirPair.From+'-'+AirPair.To)

        Capacity = {}                       # build 'capacity' constraints
        for m in range(len(AirportPairs)):
            AirPair = AirportPairs[m]
            Capacity[AirPair.From,AirPair.To] = model.addConstr(x[AirPair.From,AirPair.To]
                                                                + quicksum(w[AirPair.From,p.ICAO]*(1-next(airp for airp in Airports if airp.ICAO == AirPair.To).Hub) for p in Airports)
                                                                + quicksum(w[p.ICAO,AirPair.To]*(1-next(airp for airp in Airports if airp.ICAO == AirPair.From).Hub) for p in Airports)
                                                                ,'<=', quicksum(z[AirPair.From,AirPair.To,k]*Aircrafts[k].Seats*load_factor for k in range(len(Aircrafts))), name = 'Capacity'+AirPair.From+'-'+AirPair.To)

        Continuity = {}
        for k in range(len(Aircrafts)):                       # build 'continuity' constraints
            for m in range(len(Airports)):    
                Continuity[Airports[m]] = model.addConstr((quicksum(z[Airports[m].ICAO,p.ICAO,k] for p in Airports) - quicksum(z[p.ICAO,Airports[m].ICAO,k] for p in Airports)),
                                        '=', 0, name ='Continuity' +Airports[m].ICAO+Aircrafts[k].Name)

        #AirportPairs[m].Distance/Aircrafts[0].Speed+Aircrafts[0].LTO)*z[AirportPairs[m].From,AirportPairs[m].To]

        ACProductivity = {}                       # build 'capacity' constraints
        for k in range(len(Aircrafts)):
            ACProductivity[k] = model.addConstr(quicksum(((AirportPairs[m].Distance/Aircrafts[k].Speed+Aircrafts[k].TAT*(1+0.5*(1-next(airp for airp in Airports if airp.ICAO == AirportPairs[m].To).Hub)))*z[AirportPairs[m].From,AirportPairs[m].To,k]) for m in range(len(AirportPairs))),
                                                                    '<=', block_time*ACk[k], name = 'ACProductivity'+Aircrafts[k].Name)

        Runway = {}
        for k in range(len(Aircrafts)):
            for m in range(len(AirportPairs)):
                AirPair = AirportPairs[m]
                acft = Aircrafts[k]
                airp_from = next(airp for airp in Airports if airp.ICAO == AirPair.From).Runway
                airp_to   = next(airp for airp in Airports if airp.ICAO == AirPair.To).Runway
                if airp_from >= acft.Runway and airp_to >= acft.Runway:
                    a = 10000
                else:
                    a = 0
                Runway[AirPair.From,AirPair.To,k] = model.addConstr(z[AirPair.From,AirPair.To,k],
                                                                    '<=', a, 
                                                                    name = 'runway'+AirPair.From+'-'+AirPair.To+acft.Name)

        Range = {}
        for k in range(len(Aircrafts)):
            for m in range(len(AirportPairs)):
                AirPair = AirportPairs[m]
                if AirPair.Distance <= Aircrafts[k].Range:
                    a = 10000
                else:
                    a = 0
                Range[AirPair.From,AirPair.To,k] = model.addConstr(z[AirPair.From,AirPair.To,k],
                                                                    '<=', a, 
                                                                    name = 'range'+AirPair.From+'-'+AirPair.To+Aircrafts[k].Name)

    
    if solve:
        model.update()
        # model.setParam('Ti meLimit', 3*3600)
        model.write("FN_Model.lp")    
        model.optimize()
        model.write("FN_Model.sol")
    else:
        model.update()
        model.read("FN_Model.sol")
        model.optimize()

    status = model.status
    if status != GRB.Status.OPTIMAL:
        if status == GRB.Status.UNBOUNDED:
            print('The model cannot be solved because it is unbounded')
        elif status == GRB.Status.INFEASIBLE:
            print('The model is infeasible; computing IIS')
            model.computeIIS()
            print('\nThe following constraint(s) cannot be satisfied:')
            for c in model.getConstrs():
                if c.IISConstr:
                    print('%s' % c.constrName)
        elif status != GRB.Status.INF_OR_UNBD:
            print('Optimization was stopped with status %d' % status)
        # exit(0)

    print

    ask = 0
    for k in range(len(Aircrafts)):
            for m in range(len(AirportPairs)):
                AirPair = AirportPairs[m]
                ask += AirPair.Distance*z[AirPair.From,AirPair.To,k].X*Aircrafts[k].Seats

    print("ask",ask)
    
    rpk = 0
    for m in range(len(AirportPairs)):
        AirPair = AirportPairs[m]
        rpk += (AirPair.Distance*(x[AirPair.From,AirPair.To].X + quicksum(w[AirPair.From,p.ICAO].X*(1-next(airp for airp in Airports if airp.ICAO == AirPair.To).Hub) for p in Airports)
                + quicksum(w[p.ICAO,AirPair.To].X*(1-next(airp for airp in Airports if airp.ICAO == AirPair.From).Hub) for p in Airports)))
    print("rpk",rpk)
   
    total_paxx = 0
    for m in range(len(AirportPairs)):
        total_paxx += (x[AirportPairs[m].From,AirportPairs[m].To].X)
    print('pax_X',total_paxx)
    total_paxw = 0
    for m in range(len(AirportPairs)):
        total_paxw += ( w[AirportPairs[m].From,AirportPairs[m].To].X)
    print('pax_W',total_paxw)



    costs = 0
    for k in range(len(Aircrafts)):
        acft = Aircrafts[k]
        costs += ACk[k].X * acft.Lease
        for m in range(len(AirportPairs)):
            AirPair = AirportPairs[m]                
            gi = next(airp for airp in Airports if airp.ICAO == AirPair.From).Hub
            gj = next(airp for airp in Airports if airp.ICAO == AirPair.To).Hub
            costs += ((acft.OperatingCost + acft.TimeCost*AirPair.Distance/acft.Speed + 
                                                            acft.FuelCost*fuel_price/1.5*AirPair.Distance)*(1-(2- gi - gj)*0.3)
                                                            +energy_price*acft.EnergyCost*AirPair.Distance/acft.Range)*z[AirPair.From, AirPair.To,k].X
    print("costs",costs)

    revenue = 0
    for m in range(len(AirportPairs)):
        AirPair = AirportPairs[m]
        if AirPair.Distance > 0:
            yield_rpk = 5.9*AirPair.Distance**-0.76+0.043
        else:
            yield_rpk = 0 
        revenue += yield_rpk*AirPair.Distance*(x[AirPair.From,AirPair.To].X + 0.9 * w[AirPair.From,AirPair.To].X)
    print('revenue',revenue)
    
    # rask = float(revenue/ask)
    # Yield = float(revenue/rpk)
    # LF = float(rpk/ask)

    utilization_lst = []
    for k in range(len(Aircrafts)):
        productivity = block_time*ACk[k].X
        print('productivity',productivity)
        act_block_time = quicksum(((AirportPairs[m].Distance/Aircrafts[k].Speed+Aircrafts[k].TAT*(1+0.5*(1-next(airp for airp in Airports if airp.ICAO == AirportPairs[m].To).Hub)))*z[AirportPairs[m].From,AirportPairs[m].To,k].X) for m in range(len(AirportPairs)))
        print("act_block_time",act_block_time)
        # utilization_lst.append(act_block_time / productivity)

    # print("ask",ask)
    # print("rpk",rpk)
    # print("cask",cask)
    # print("rask",rask)
    # print("yield",Yield)
    # print('lf',LF)
    # for utilization in utilization_lst:
    #     print('utilization',utilization)
              
                
    #     for k in range(len(Aircrafts)):
    #         if z[AirportPairs[m].From,AirportPairs[m].To,k].X >0:
    #             print("z"+AirportPairs[m].From+'-'+AirportPairs[m].To+'-'+Aircrafts[k].Name,z[AirportPairs[m].From,AirportPairs[m].To,k].X)
    
    # for k in range(len(Aircrafts)):
    #     if ACk[k].X > 0:
    #         print("AC"+Aircrafts[k].Name,ACk[k].X)

    workbook = xlsxwriter.Workbook('results.xlsx')
    worksheet = workbook.add_worksheet()

   

    cell_format1 = workbook.add_format()
    cell_format1.set_num_format(9)
    cell_format1.set_left()

    cell_format2 = workbook.add_format()
    cell_format2.set_top()

    cell_format3 = workbook.add_format()
    cell_format3.set_left()

    cell_format4 = workbook.add_format()
    cell_format4.set_left()
    cell_format4.set_top()


    for col in range(len(Airports)):
        worksheet.write(0, col*2+1, Airports[col].ICAO)
        worksheet.write(col*3+1, 0, Airports[col].ICAO)
        for row in range(len(Airports)):
            From    = Airports[col].ICAO
            To      = Airports[row].ICAO
            worksheet.write((row)*3+1,(col)*2+1,x[From,To].X+w[From,To].X,cell_format4)            
            worksheet.write((row)*3+2,(col)*2+1,w[From,To].X,cell_format3)            
            if next(airp for airp in AirportPairs if airp.From == From and airp.To == To).Demand2030 == 0:
                worksheet.write((row)*3+3,(col)*2+1,0,cell_format1)
            else:
                worksheet.write((row)*3+3,(col)*2+1,(x[From,To].X+w[From,To].X)/next(airp for airp in AirportPairs if airp.From == From and airp.To == To).Demand2030,cell_format1)
            worksheet.write((row)*3+1,(col)*2+2,z[From,To,0].X,cell_format2)
            worksheet.write((row)*3+2,(col)*2+2,z[From,To,1].X)
            worksheet.write((row)*3+3,(col)*2+2,z[From,To,2].X)    

    workbook.close()
            
    print
    print ("Objective Function =", model.ObjVal/1.0)
    print ("------------------------------------------------------------------------")


if __name__ == '__main__':
    #=================================================================================================
    # Input excel file with arcs data (sheet1) and commodities data (sheet2)
    fuel_price      = 1.42  # $/gallon
    block_time      = 10*7    # hrs
    energy_price    = 0.07  # euro/kWh
    load_factor     = 0.8
    AirportPairs    = []
    Aircrafts       = []
    Airports        = []
    DemandPairs     = []

    wb = load_workbook("Aircraft_info.xlsx", read_only=True)
    List_aircraft_info = tuple(wb["Aircraft_info"].iter_rows())

    wb = load_workbook("Group_16_Airport_info.xlsx", read_only=True)
    List_airport_info = tuple(wb["Group_16_Airport_info"].iter_rows())

    wb = load_workbook("Group_16_Demand.xlsx", read_only=True)
    List_demand_forecast_data = tuple(wb["Group_16_Demand"].iter_rows())

    wb = load_workbook("Group_16_Distances.xlsx", read_only=True)
    List_airport_distances = tuple(wb["Group_16_Distances"].iter_rows())

    wb = load_workbook("Group_16_Annual_Growth.xlsx", read_only=True)
    List_annual_growth = tuple(wb["Group_16_Annual_growth"].iter_rows())

    annual_growth = float(List_annual_growth[0][0].value)

    # origin, destination, distance, demand2020, demand2030
    for i, origin in enumerate(List_demand_forecast_data[1:]):
        for j, destination in enumerate(origin[1:]):
            new_pair = DemandAirportPair(List_demand_forecast_data[i+1][0].value,List_demand_forecast_data[0][j+1].value,int(List_demand_forecast_data[i+1][j+1].value))
            DemandPairs.append(new_pair)
    
    # name, icao, latitude, longitude, runway_length, population, gdp, hub
    for (name, icao, latitude, longitude, runway_length, population, gdp, hub) in List_airport_info[1:]:
        new = Airport(name.value, icao.value, latitude.value, longitude.value, int(runway_length.value), int(population.value), None, gdp.value, int(hub.value))
        Airports.append(new)

    for i, origin in enumerate(List_airport_distances[1:]):
        for j, destination in enumerate(origin[1:]):
            new_pair = AirportPair(List_airport_distances[i+1][0].value,List_airport_distances[0][j+1].value,(List_airport_distances[i+1][j+1].value), None)
            AirportPairs.append(new_pair)

    # for pair in AirportPairs:
    #     print(vars(pair))
    # print("")    
    # for pair in DemandPairs:
    #     print(vars(pair))

    
    X = np.zeros((len(DemandPairs),4))
    
    for i in range(len(DemandPairs)):
        for j in range(4):
            if j == 0:
                X[i,j] = 1
            elif j == 1:
                # next(airp for airp in Airports if airp.Name == AirportPairs[m].From).Population
                X[i,j] = log(next(airp for airp in Airports if airp.ICAO == DemandPairs[i].From).Population2020*
                            next(airp for airp in Airports if airp.ICAO == DemandPairs[i].To).Population2020)  
            elif j == 2:
                X[i,j] = log(next(airp for airp in Airports if airp.ICAO == DemandPairs[i].From).GDP*
                            next(airp for airp in Airports if airp.ICAO == DemandPairs[i].To).GDP)  
            elif j == 3:
                X[i,j] = log(next(airppair for airppair in AirportPairs if (airppair.From == DemandPairs[i].From and airppair.To == DemandPairs[i].To)).Distance*
                            fuel_price) 

    X = np.delete(X, np.where(X < 0)[0], axis=0)

    y = np.zeros(len(DemandPairs))
    for i in range(len(y)):
        y[i] = log(DemandPairs[i].Demand2020)

    y = np.delete(y, np.where(y < 0)[0], axis=0)

    solution = np.linalg.lstsq(X,y)[0]

    k = solution[0]
    b1 = solution[1]
    b2 = solution[2]
    b3 = solution[3]

    k = -5.908334491674755
    b1 = 0.3507381298874774
    b2 = 0.14095537035529815
    b3 = -0.25329734075161936

    # print("k = " ,solution[0])
    # print("b1 = ",solution[1])
    # print("b2 = ",solution[2])
    # print("b3 = ",solution[3])

    for airport in Airports:
        airport.Population2030 = airport.Population2020*(annual_growth**10)

    for airport_pair in AirportPairs:
        if airport_pair.From != airport_pair.To:
            population_term = b1*log(next(airp for airp in Airports if (airp.ICAO == airport_pair.From)).Population2030
                                                *next(airp for airp in Airports if (airp.ICAO == airport_pair.To)).Population2030)
            gdp_term = b2*log(next(airp for airp in Airports if (airp.ICAO == airport_pair.From)).GDP
                                                *next(airp for airp in Airports if (airp.ICAO == airport_pair.To)).GDP)
            fuel_term = b3*log(fuel_price*airport_pair.Distance)

            airport_pair.Demand2030 = exp(k + population_term + gdp_term + fuel_term)
        else:
            airport_pair.Demand2030 = 0

    for (name, speed, seats, tat, charging, ac_range, runway, lease_cost, op_cost, time_cost, fuel_cost, energy_cost) in List_aircraft_info[1:4]:
        new = Aircraft(name.value, speed.value, seats.value, tat.value, charging.value, ac_range.value, runway.value, lease_cost.value, op_cost.value, time_cost.value, fuel_cost.value, energy_cost.value)

        Aircrafts.append(new)
           
    
    # for pair in AirportPairs:
        
    #     print(pair.From+pair.To)
    #     print(pair.Demand2030)
    #     print('')

    # D = np.zeros((len(Airports),len(Airports)))


    # for i in range(len(Airports)):
    #     for j in range(len(Airports)):
    #         D[i,j] = next(airp for airp in AirportPairs if airp.From == Airports[i].ICAO and airp.To == Airports[j].ICAO).Demand2030
    
    # print(D)
    # pd.DataFrame(D).to_csv("testtt.csv")


    # for pair in Airports:
    #     print(vars(pair))


    start_time = time()
    
    FN_Problem(AirportPairs, Aircrafts, Airports, fuel_price, block_time, energy_price, load_factor, solve=False)
    
    elapsed_time = time() - start_time

    print ("Run Time = ", elapsed_time)
    print ("END")



    

 

    